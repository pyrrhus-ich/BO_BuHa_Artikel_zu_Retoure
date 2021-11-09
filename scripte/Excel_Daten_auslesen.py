import os
import openpyxl as op

print("Es geht los")
# <<< Dieser Bereich definiert die Variablen für das gesamte Script >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
workDir = os.getcwd()                   # legt das Arbeitsverzeichnis fest
srcFldBuHa=workDir + "/src_BuHa/"              #Gibt den Ordner der Source Files an. Hier liegt das File aus der BuHa
srcFldBo = workDir + "/src_BO/"             #Hier liegt das File mit dem BO_Bericht
srcFileBuHa = srcFldBuHa + os.listdir(srcFldBuHa)[0]      #Liest den Namen des BuHa Files
srcFileBo = srcFldBo + os.listdir(srcFldBo)[0]     #Liest den Namen des Bo Berichtes

#Hier werden der Pfad und der Filename des Zielfiles angegeben
dstFileName="ArtikelZuRetour.xlsx"
dstFile='./dst/'+dstFileName

#<<<< Hier beginnt die Arbeit mit dem Buchaltungsfile >>>>>>
wbuHa = op.load_workbook(srcFileBuHa,data_only=True) # lädt das File
wsbuHa = wbuHa.worksheets[0]
#schreibt die Werte in eine Liste wenn der erste value der Zeile nicht none ist
buHaValList=[];
for value in wsbuHa.iter_rows(min_row=1, values_only=True):
        if value[0] is not None:
            buHaValList.append(value)
# Ab hier fangen wir an die Retourennummer zusammen zu bauen im Form M199-122334
buHaEindeutigeRetourennummer =[]
for el in buHaValList:
    buHaEindeutigeRetourennummer.append(el[0]+"-"+ str(el[1]))
# Die zusammengesetzte 'Eindeutige Retourennummer' wird in die Spalte C geschrieben
buHaIndEl = 1
for el in buHaEindeutigeRetourennummer:
    zelle = "C" + str(buHaIndEl)
    wsbuHa[zelle] = el 
    buHaIndEl+=1
#Nachdem in Spalte 'C' die korrekte zusammengesetzte Retourennummer steht löschen wir Spalten 'A' und 'B'
wsbuHa.delete_cols(1,2)
# Spaltenbreite wird manuell festgesetzt
wsbuHa.column_dimensions['A'].width = 16
# Bericht wird gespeichert
wbuHa.save(srcFileBuHa)
#<<< Das Source File der Buchhaltung ist fertig bearbeitet

#<<< Hier beginnt die Bearbeitung des BO Berichtes <<<<
wb = op.load_workbook(srcFileBo,data_only=True) # lädt das File
ws = wb.worksheets[0] # definiert das worksheet mit dem wir arbeiten
#ws = wb['BA_BO_Bericht']; # definiert das worksheet mit dem wir arbeiten

#schreibt die Werte in eine Liste wenn der erste value der Zeile nicht none ist
valList=[];
for value in ws.iter_rows(min_row=1, values_only=True):
        if value[0] is not None:
            valList.append(value)

#Füge eine Spalte in das Excel File ein. Dort kommt die zusammengesetzte Retourennummer rein
ws.insert_cols(3)
# Speicher den Bericht zwischen
wb.save(srcFileBo)
# Ab hier fangen wir an die Retourennummer zusammen zu bauen im Form M199-122334
eindeutigeRetourennummer =[]
for el in valList:
    eindeutigeRetourennummer.append(el[0]+"-"+ str(el[1]))
# Die zusammengesetzte 'Eindeutige Retourennummer' wird in die Spalte C geschrieben
indEl = 1
for el in eindeutigeRetourennummer:
    zelle = "C" + str(indEl)
    ws[zelle] = el 
    indEl+=1
#Nachdem in Spalte 'C' die korrekte zusammengesetzte Retourennummer steht löschen wir Spalten 'A' und 'B'
ws.delete_cols(1,2)
# Spaltenbreite wird manuell festgesetzt
ws.column_dimensions['A'].width = 16
# Bericht wird gespeichert
wb.save(srcFileBo)
#>>> BO Bericht ist geschrieben und enthält jetzt eine Spalte C mit der zusammengesetzten Retourennummer <<<

#<<< Die Vorbereitungen sind abgeschlossen, wir beginne mit dem Vergleich der zusammengesetzten Retourennummern
# Wir verfügen ja schon über die Liste mit den zusammengesetzen BuHa Nummern. "buHaEindeutigeRetourennummer"
# Jetzt müssen die Werte in dieser Liste verglichen werden mit den Werten in der Bo Tabelle
#Dazu lesen wir das Bo Source file nochmal aus
wb = op.load_workbook(srcFileBo,data_only=True) # lädt das File
ws = wb.worksheets[0] # definiert das worksheet mit dem wir arbeiten

#buHaEindeutigeRetourennummer=["M116-190550181","M187-190448059"]
valList=[];
for value in ws.iter_rows(min_row=1, values_only=True):
        if value[0] is not None:
            valList.append(value)

#Jetzt haben wir 2 Listen mit Werten "buHaEindeutigeRetourennummer" und "valList" Wenn jetzt ein Wert aus "buHaEindeutigeRetourennummer"
#mit dem Index [n][0] übereinstimmt, übergeben wir die ganze Reihe an eine neue Liste "resultList"

resultList=[]
for el in buHaEindeutigeRetourennummer:
    for val in valList:
        if val[0]==el:
            resultList.append(val)

#Die "resultListe" enthält jetzt alle Zeilen aus dem BO Bericht die auch in der Buha vorhanden sind
# Der nächste Schritt ist das Schreiben des neuen Files

#Erzeugt das neue File 
wbDst = op.Workbook() # erzeugt Workbook Objekt
ws = wbDst.active     # erzeugt das erste sheet

# <<<<<<<<< PROBLEM >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
IndWs = 0  #Zähler für die for schleife
for el in resultList:
    ws['A'+str(IndWs+1)] = resultList[IndWs]
    IndWs+=1
# <<<<<<<<<< PROBLEM ENDE >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

wbDst.save(dstFile) #Speichert das File
wb.close()     
print("\nFertig")
    

