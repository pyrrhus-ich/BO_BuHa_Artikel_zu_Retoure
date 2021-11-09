import openpyxl as op
import os
from openpyxl.descriptors.base import Length
from openpyxl.worksheet.dimensions import RowDimension
from openpyxl.worksheet.table import Table, TableStyleInfo

workDir = os.getcwd()                   # legt das Arbeitsverzeichnis fest
srcFldBuHa=workDir + "/src_BuHa/"              #Gibt den Ordner der Source Files an Hier liegt das File aus der BuHa
srcFldBo = workDir + "/src_BO/"             #Hier liegt das File mit dem BO_Bericht
srcFileBuHa = srcFldBuHa + os.listdir(srcFldBuHa)[0]      #Liest den Namen des BuHa Files
srcFileBo = srcFldBo + os.listdir(srcFldBo)[0]     #Liest den Namen des Bo Berichtes

#Hier werden der Pfad und der Filename des Zielfiles angegeben
dstFileName="ArtikelZuRetour.xlsx"
dstFile='./dst/'+dstFileName

'''
#Gibt die Verzeichnisse und Dateinamen aus
print("Arbeitsverzeichnis : " + workDir + 
"\nSource Folder BO : " + srcFldBo +
"\nSource File BO : "+ srcFileBo +
"\nSource Folder Buchhaltung : " + srcFldBuHa + 
"\nSource File Buchhaltung : " + srcFileBuHa) 
'''

wb = op.load_workbook(srcFileBuHa,data_only=True) # lädt das File

#print(wb.worksheets)        # Liste von Objekten
#print(wb.worksheets[0])
ws = wb.worksheets[0]
#print(ws.columns)
'''
for row in ws.values:  # Gibt alle Zeilen als Tuple aus
    print(row)

for column in ws.values:
    print(column)
'''
#spalten = str(ws.max_column); 
# print("Anzahl der Spalten : "+ spalten)  # Anzahl der Spalten : n
#zeilen = str(ws.max_row)
#print("Anzahl der Zeilen " + zeilen)     # Anzahl Zeilen : n

#liest die einzelnen Zeilen aus und hängt sie an die Liste
retListe = list();

for values in ws.iter_rows(values_only=True):
    val = values[0] + "-" + str(values[1])
    retListe.append(values[0] + "-" + str(values[1]));
    
print("retList : " + str(retListe))

print("Erfolg") if len(retListe)==ws.max_row else print("Fehler"); # Prüft ob die Elemente in der Liste die gleiche Anzahl haben wie die
                                                                   # maximale anzahl an Zeilen

#Erzeugt das neue File 
wbDst = op.Workbook() # erzeugt Workbook Objekt
ws = wbDst.active     # erzeugt das erste sheet

#schreibt alle Werte aus 'retListe' in das Sheet

IndWs = 0  #Zähler für die for schleife
for el in retListe:
    ws['A'+str(IndWs+1)] = retListe[IndWs]
    IndWs+=1

#Spaltenbreite anpassen vorläufig manuell. Das kann man aber auch berechnen lassen
ws.column_dimensions['A'].width = 16

wbDst.save('ziel.xlsx') #erzeugt das File
wb.close()


    





