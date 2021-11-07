import openpyxl as op
from settings import *

wb = op.load_workbook(srcFile,data_only=True) # lädt das File
print(wb.sheetnames)        #druckt alle Arbeitsblattnamen als Liste von Strings
print(wb.sheetnames[0])     # gibt den Namen an Index 0 aus
sn = wb.sheetnames          # Speichert die Liste der Namen
print(sn[0])                # gibt index 0 aus
for el in sn:               # iteration über alle Blattname
    print("sheetnames " + el)
print(wb.worksheets)        # Liste von Objekten
print(wb.worksheets[0])




