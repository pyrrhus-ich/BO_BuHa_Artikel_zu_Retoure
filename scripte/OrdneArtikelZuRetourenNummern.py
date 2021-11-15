import os, csv
from Basis import folderList
from LegeVerzeichnisseNeuAn import checkFolders
import openpyxl as op
from openpyxl.styles import PatternFill, Font


'''
- Diese Script nimmt ein CSV File(Bo Bericht) aus dem Ordner Excel, wandelt es um in ein XSLX File und packt es nach 'src_BO
- Im Ordner src_BuHa befindet sich ein Excel File mit Markt und Retourennummer Spalte A ist Marktnummer (M199) Spalte B ist die Retourennummer
- Im dst Ordner findet sich dann das Endergebniss
- Wegen GIT sind alle Files aus den Ordnern entfernt
Vorraussetzungen:
    Ordnerstruktur:
        Excel---|
                |-dst
                |-src_BuHa
                |-src_BO
                |-scripte
                |-logfiles
    1. BO Bericht aus BO\Meine Favoriten\96_BuHa...\BA_BO_Bericht ist erstellt und als CSV exportiert
    2. Das erstellte CSV File liegt im Stammverzeichniss (Hier also 'Excel')
    3. Das File der Buchhaltung ist ein Excel File mit 2 Spalten (Spalte A ist Marktnummer (M199) Spalte B ist die Retourennummer)
Wenn alle Vorraussetzungen erfüllt sind, sollte es funktionieren.
'''

print("Es geht los")

checkFolders(folderList);


'''
# <<< Dieser Bereich definiert die Variablen für das gesamte Script >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

workDir = os.getcwd()                                   # legt das Arbeitsverzeichnis fest
csvSrc = "BA_BO_Bericht.csv"                            # Gibt das source CSV File an
srcFldBuHa=workDir + "/src_BuHa/"                       #Gibt den Ordner der Source Files an. Hier liegt das File aus der BuHa
srcFldBo = workDir + "/src_BO/"                         #Hier liegt das File mit dem BO_Bericht
srcFileBuHa = srcFldBuHa + os.listdir(srcFldBuHa)[0]    #Liest den Namen des BuHa Files
srcFileBo =''                                           # wird erst mal nur deklariert
#Hier werden der Pfad und der Filename des Zielfiles angegeben
dstFileName="Zuordnung_Artikel_zur_Retoure.xlsx"
dstFile='./dst/'+dstFileName

print("Variablen sind deklariert")

#<<<<<<< Hier wird das CSV File in ein XLSX File umgewandelt >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
print("csv wird in Excel umgewandelt")
csvWb = op.Workbook()
csvWs = csvWb.active

with open(csvSrc) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        csvWs.append(row)

csvWb.save(srcFldBo + 'BA_BO_Bericht.xlsx')
f.close()
print("Excel File wurde erstellt in : " + srcFldBo)
#<<<<<<<<<<<<<<< Excel File sollte erstellt sein >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
#<<<<<<<<<Letzte Variable wird erstellt >>>>>>>>>>
srcFileBo = srcFldBo + os.listdir(srcFldBo)[0]          #Liest den Namen des Bo Berichtes


#<<<< Hier beginnt die Arbeit mit dem Buchaltungsfile >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
print("Auslesen BuHa src File beginnt")
wbuHa = op.load_workbook(srcFileBuHa,data_only=True) # lädt das File
wsbuHa = wbuHa.worksheets[0]
#schreibt die Werte in eine Liste wenn der erste value der Zeile nicht none ist
buHaValList=[];
for value in wsbuHa.iter_rows(min_row=1, values_only=True):
        if value[0] is not None:
            buHaValList.append(value)
# Ab hier fangen wir an die Retourennummer zusammen zu bauen im Form M199-122334
buHaEindeutigeRetourennummer =[]
for el in buHaValList:
    buHaEindeutigeRetourennummer.append(el[0]+"-"+ str(el[1]))

print("Das auslesen des BuH src Files ist beendet")
#<<< Das Source File der Buchhaltung ist fertig bearbeitet >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

#<<< Hier beginnt die Bearbeitung des BO Berichtes >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
print("Bearbeitung des BO Files beginnt")

wb = op.load_workbook(srcFileBo,data_only=True) # lädt das File
ws = wb.worksheets[0] # definiert das worksheet mit dem wir arbeiten

#schreibt die Werte in eine Liste wenn der erste value der Zeile nicht none ist
valList=[];
for value in ws.iter_rows(min_row=1, values_only=True):
        if value[0] is not None:
            valList.append(value)

#Füge eine Spalte in das Excel File ein. Dort kommt die zusammengesetzte Retourennummer rein
ws.insert_cols(3)
# Speicher den Bericht zwischen
wb.save(srcFileBo)
# Ab hier fangen wir an die Retourennummer zusammen zu bauen im Form M199-122334
eindeutigeRetourennummer =[]
for el in valList:
    eindeutigeRetourennummer.append(el[0]+"-"+ str(el[1]))
# Die zusammengesetzte 'Eindeutige Retourennummer' wird in die Spalte C geschrieben
indEl = 1
for el in eindeutigeRetourennummer:
    zelle = "C" + str(indEl)
    ws[zelle] = el 
    indEl+=1

# Bericht wird gespeichert
wb.save(srcFileBo)

print("BO File ist fertig bearbeitet")
#>>> BO Bericht ist geschrieben und enthält jetzt eine Spalte C mit der zusammengesetzten Retourennummer <<<

#<<< Die Vorbereitungen sind abgeschlossen, wir beginne mit dem Vergleich der zusammengesetzten Retourennummern
# Wir verfügen ja schon über die Liste mit den zusammengesetzen BuHa Nummern. "buHaEindeutigeRetourennummer"
# Jetzt müssen die Werte in dieser Liste verglichen werden mit den Werten in der Bo Tabelle
#Dazu lesen wir das Bo Source file nochmal aus
print("Starte den Vergleich der beiden Listen" )
wb = op.load_workbook(srcFileBo,data_only=True) # lädt das File
ws = wb.worksheets[0] # definiert das worksheet mit dem wir arbeiten

valList=[];
for value in ws.iter_rows(min_row=1, values_only=True):
        if value[0] is not None:
            valList.append(value)

#Jetzt haben wir 2 Listen mit Werten: "buHaEindeutigeRetourennummer" und "valList" Wenn jetzt ein Wert aus "buHaEindeutigeRetourennummer"
#mit dem Index [n][0] der 'valList' übereinstimmt, übergeben wir die ganze Reihe an eine neue Liste "resultList"

resultList=[("BuKr","Zuordnung","Zusammengesetzt","Artikelnummer","Artikelbezeichnung")];

for el in buHaEindeutigeRetourennummer:
    for val in valList:
        if val[2]==el:
            resultList.append(val)

print("Der Vergleich der beiden Listen ist abgeschlossen")
#Die "resultListe" enthält jetzt alle Zeilen aus dem BO Bericht die auch in der Buha vorhanden sind
# Der nächste Schritt ist das Schreiben des neuen Files
print("Beginne jetzt mit dem schreiben des dst Files")
#Erzeugt das neue File 
wbDst = op.Workbook() # erzeugt Workbook Objekt mit einem Sheet
ws = wbDst.active     # aktiviert das erste sheet

for row in resultList:
    ws.append(row)
#<<<<<<<<<< Hier beginnt die Formatierung des Excel Files >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
print("Wir starten das schön machen des Files")
#Löscht die Spalte C mit der zusammengesetzten Nummer
ws.delete_cols(3)
# Setzt die Spaltenbreiten
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 100

#Setzt den Autofilter über alle Spalten
ws.auto_filter.ref = ws.dimensions

# Setze für die Überschrift den Background Color zu Blau | Schriftgröße zu 14 | Schriftfarbe zu Weiss
for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=4):
    for cell in rows:
      cell.fill = PatternFill(start_color="2f48ff", end_color="2f48ff", fill_type = "solid")
      cell.font = Font(size=14, color="FFFFFF")




#ws['A1'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")
print("Alles ist schön. File kann gespeichert werden")
wbDst.save(dstFile) #Speichert das File
wbDst.close()
print("Das schreiben des Ausgabefiles ist abgeschlossen. Es wird bereitgestellt in : " + dstFile)   
print("\nFertig Alles erledigt. Das File kann versendet werden")

'''
