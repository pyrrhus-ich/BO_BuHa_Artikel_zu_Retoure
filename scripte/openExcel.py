import openpyxl as op
from settings import *

wb = op.load_workbook(srcFileBuHa)
ws = wb.sheetnames[0]
# print(wb.sheetnames) # ['Tabelle1']
# print(ws) # Tabelle1



