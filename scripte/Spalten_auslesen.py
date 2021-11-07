import openpyxl as op
from openpyxl.descriptors.base import Length
from openpyxl.worksheet.dimensions import RowDimension
from settings import *

wb = op.load_workbook(srcFileBuHa,data_only=True) # lädt das File

#print(wb.worksheets)        # Liste von Objekten
#print(wb.worksheets[0])
ws = wb.worksheets[0]
#print(ws.columns)
'''
for row in ws.values:  # Gibt alle Zeilen als Tuple aus
    print(row)

for column in ws.values:
    print(column)
'''
#spalten = str(ws.max_column);
# print("Anzahl der Spalten : "+ spalten)  # Anzahl der Spalten : n
#zeilen = str(ws.max_row)
#print("Anzahl der Zeilen " + zeilen)     # Anzahl Zeilen : n

#liest die einzelnen Zeilen aus und hängt sie an die Liste
retListe = list();
retTuple = ();
for values in ws.iter_rows(values_only=True):
    val = values[0] + "-" + str(values[1])
    retListe.append(values[0] + "-" + str(values[1]));
    retTuple = retTuple + (str(val),)
    
print(retTuple)
#prüft ob alle Zeilen gelesen wurden    
print("Erfolg") if len(retListe)==ws.max_row else print("Fehler");




    





