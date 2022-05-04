import os
import openpyxl as op
from openpyxl.styles import PatternFill, Font
from colorama import init, Fore, Style # Colorama für farbige Terminalausgaben
#init()

# Definiert die Variablen

srcLst = []
dstLst = []
wrgLst = []
dstEl = []

workDir = os.getcwd()           
srcFld=workDir + "\\srcZir\\"    
srcFile = srcFld + os.listdir(srcFld)[0]  
dstFld=workDir + "\\dstZir\\"
dstFile= dstFld + "Zieldatei.xlsx"


# liest das src File und speichert die Werte als Liste in einer Variablen ab
def readSrc(sourceFile, dstLstSrc):
    print("Auslesen src File beginnt : readSrc()")
    wb = op.load_workbook(sourceFile,data_only=True) # lädt das File
    ws = wb.worksheets[0]
    #schreibt die Werte in eine Liste wenn der erste value der Zeile nicht none ist
    for value in ws.iter_rows(min_row=2, values_only=True):
            if value[1] is not None:
                dstLstSrc.append(value)
                #print(value)
                #print("")
    print("srcFile wurde gelesen\n")

# Erstellt die zukünftige Kopfzeile als erstes Element in der Ziellist (dstList)
# Je nach Bedarf muss das noch entsprechend angepasst werden
# liEl = Das einzufügende Element
# dst = die Zielliste(dstLst) in die liEl eingefügt wird
def crtHeadline(liEl, dst):
    liEl.append("HSN Nr.")
    liEl.append("Hersteller")
    liEl.append("WGR")
    liEl.append("WGR Name")
    dst.append(liEl)

# Erstellt alle weiteren Elemente der Zielliste (dstLst)
# Hier erfolgt dann die weitere Bearbeitung der einzelnen Listelemente der Ursprungsliste
# In diesem Beispiel ging es darum die Warengruppennummern die jeweils in i[3] der Sourcelistelement gespeichert sind
# so zu splitten, das man darüber iterieren kann und dann für jede Warengruppe eine Zeile bekommt
def crtDstLst(src, liEl, dst):
    liEl = []
    for i in src:
        cnt = 0
        id = i[3].split(",")
        id = list(set(id)) # Entfernt Doubletten
        while cnt < len(id):
                liEl.append(i[0])
                liEl.append(i[1])
                liEl.append(id[cnt])
                cnt += 1
                #print(liEl)
                dst.append(liEl)
                liEl = []
                print(dst)

# Hier erstellen wir das endgültige Excelfile
def createDstFile(baseList, dstFileName): #resulList, dstFile
    print("Beginne jetzt mit dem schreiben des dst Files createDstFile()")
    wbDst = op.Workbook() # erzeugt Workbook Objekt mit einem Sheet
    ws = wbDst.active     # aktiviert das erste sheet
    for row in baseList: #Hängt jeden Eintrag in der resultList an das neu erzeugt sheet
        ws.append(row)
    #<<<<<<<<<< Hier beginnt die Formatierung des Excel Files >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    #print("Wir starten das schön machen des Files")
    # Setzt die Spaltenbreiten
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 50
    #Setzt den Autofilter über alle Spalten
    ws.auto_filter.ref = ws.dimensions          
    # Setze für die Überschrift den Background Color zu Blau | Schriftgröße zu 14 | Schriftfarbe zu Weiss
    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=4):
        for cell in rows:
            cell.fill = PatternFill(start_color="2f48ff", end_color="2f48ff", fill_type = "solid")
            cell.font = Font(size=14, color="FFFFFF")   
    wbDst.save(dstFileName) #Speichert das File
    wbDst.close()
    print(Fore.GREEN + "Das schreiben des Ausgabefiles ist abgeschlossen. Es wird bereitgestellt in : " + dstFileName)   
    print(Style.RESET_ALL, end="")#Setzt die Farbeinstellungen wieder zurück
    #print("\nFertig Alles erledigt. Das File kann versendet werden")


    
