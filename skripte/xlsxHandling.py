import openpyxl as op
from openpyxl.styles import PatternFill, Font
from colorama import init, Fore, Style # Colorama für farbige Terminalausgaben
from skripte.basis import buHaEindeutigeRetourennummer, valList, buHaValList, nichtGefundeneRetouren
init()




# liest das BuHa File und setzt die Werte in einer Liste zusammen. Dadurch entsteht die Eindeutige Retourennummer die sich dann mit dem BO Bericht vergleichen lässt
# dies ist erforderlich, da unsere Retourennummern über alle Märkte hinweg nicht eineindeutig sind, sondern doppelt vorkommen können.
def readSrcBuHa(source):
    print("Auslesen BuHa src File beginnt : readSrcBuHa()")
    wbuHa = op.load_workbook(source,data_only=True) # lädt das File
    wsbuHa = wbuHa.worksheets[0]
    #schreibt die Werte in eine Liste wenn der erste value der Zeile nicht none ist
    for value in wsbuHa.iter_rows(min_row=1, values_only=True):
            if value[0] is not None:
                buHaValList.append(value)
    # Ab hier fangen wir an die Retourennummer zusammen zu bauen im Form M199-122334
    for el in buHaValList:
        buHaEindeutigeRetourennummer.append(el[0]+"-"+ str(el[1])) #Speichert die zusammengesetzten Werte in einer Liste
    #print("Das auslesen des BuH src Files ist beendet.\nDie Werte für die eindeutige Retourennummer wurden als Liste gespeichert")


#<<< Hier beginnt die Bearbeitung des BO Berichtes >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
def bearbBOXlsx(sourceFile):
    print("Bearbeitung des BO Files beginnt : bearbBOXlsx()")
    wb = op.load_workbook(sourceFile,data_only=True) # lädt das File
    ws = wb.worksheets[0] # definiert das worksheet mit dem wir arbeiten
    #schreibt die Werte in eine Liste wenn der erste value der Zeile nicht none ist
    valList=[];
    for value in ws.iter_rows(min_row=1, values_only=True):
            if value[0] is not None:
                valList.append(value)
    #Füge eine Spalte in das Excel File ein. Dort kommt die zusammengesetzte Retourennummer rein
    ws.insert_cols(3)
    # Speicher den Bericht zwischen
    wb.save(sourceFile)
    # Ab hier fangen wir an die Retourennummer zusammen zu bauen im Form M199-122334
    eindeutigeRetourennummer =[]
    for el in valList:
        eindeutigeRetourennummer.append(el[0]+"-"+ str(el[1]))
    # Die zusammengesetzte 'Eindeutige Retourennummer' wird in die Spalte C geschrieben
    indEl = 1
    for el in eindeutigeRetourennummer:
        zelle = "C" + str(indEl)
        ws[zelle] = el 
        indEl+=1
    # Bericht wird gespeichert
    wb.save(sourceFile)
    #print("BO File ist fertig bearbeitet")
#>>> BO Bericht ist geschrieben und enthält jetzt eine Spalte C mit der zusammengesetzten Retourennummer <<<

#Erstellt die Liste mit den zusammengesetzten Retourennummern aus dem BO xlsx. Diese Liste wird im Anschluss mit der BuHa Liste verglichen.
#Dabei ist wichtig zu wissen, das die einzelnen Werte der Liste bestehen aus: MarktNr, Retourennummer, ArtNr, Art Text
def createValList(sourceFile):
    wb = op.load_workbook(sourceFile,data_only=True) # lädt das File
    ws = wb.worksheets[0] # definiert das worksheet mit dem wir arbeiten
    for value in ws.iter_rows(min_row=1, values_only=True):
            if value[0] is not None:
                valList.append(value)

def createDstFile(baseList, dstFileName): #resulList, dstFile
    print("Beginne jetzt mit dem schreiben des dst Files createDstFile()")
    wbDst = op.Workbook() # erzeugt Workbook Objekt mit einem Sheet
    ws = wbDst.active     # aktiviert das erste sheet
    for row in baseList: #Hängt jeden Eintrag in der resultList an das neu erzeugt sheet
        ws.append(row)
    #<<<<<<<<<< Hier beginnt die Formatierung des Excel Files >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    #print("Wir starten das schön machen des Files")
    #Löscht die Spalte C mit der zusammengesetzten Nummer
    ws.delete_cols(3)
    # Setzt die Spaltenbreiten
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 35
    ws.auto_filter.ref = ws.dimensions          #Setzt den Autofilter über alle Spalten
    if len(nichtGefundeneRetouren)>0:           #Wenn es Retouren gibt die nicht zuzuordnen sind
        ws.column_dimensions['G'].width = 35
        ws['G1']="Nicht gefunden"
        # Setze für die Überschrift den Background Color zu Blau | Schriftgröße zu 14 | Schriftfarbe zu Weiss
        for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=7):
            for cell in rows:
                cell.fill = PatternFill(start_color="2f48ff", end_color="2f48ff", fill_type = "solid")
                cell.font = Font(size=14, color="FFFFFF")
        # schreibt die Werte 
        rangeEnd = len(nichtGefundeneRetouren)+2  # Konstrukt ist erforderlich weil er erst in Zeile 2 anfangen soll zu schreiben.
        for i in range(2,rangeEnd):
            el=i-2      # entspricht zu Beginn dem Index 0
            cnt = 2
            while cnt <= rangeEnd:
                cellref=ws.cell(row=i, column=7)  # Am Start : Zeile 2 in Spalte 7
                cellref.value=nichtGefundeneRetouren[el] 
                cnt+=1
    else:       # Sollte wirklich alle Retouren gefunden worden sein
        # Setze für die Überschrift den Background Color zu Blau | Schriftgröße zu 14 | Schriftfarbe zu Weiss
        for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=6):
            for cell in rows:
                cell.fill = PatternFill(start_color="2f48ff", end_color="2f48ff", fill_type = "solid")
                cell.font = Font(size=14, color="FFFFFF")
    wbDst.save(dstFileName) #Speichert das File
    wbDst.close()
    print(Fore.BLUE + "Das schreiben des Ausgabefiles ist abgeschlossen. Es wird bereitgestellt in : " + dstFileName)   
    print(Style.RESET_ALL, end="")#Setzt die Farbeinstellungen wieder zurück
    #print("\nFertig Alles erledigt. Das File kann versendet werden")


    
